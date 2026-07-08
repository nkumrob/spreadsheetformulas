#!/usr/bin/env python3
"""Generate the 5 launch templates (ticket SF-087) as real .xlsx files with
working formulas, into public/templates/. Idempotent — safe to re-run."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "public" / "templates"
OUT.mkdir(parents=True, exist_ok=True)

INK = "1C1A16"
LEDGER = "166B4A"
CREAM = "F3EEE4"

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=LEDGER)
SUMMARY_FONT = Font(name="Calibri", bold=True, color=INK, size=11)
SUMMARY_FILL = PatternFill("solid", fgColor=CREAM)


def style_sheet(ws, headers, widths, n_data_rows):
    for col, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def summary_block(ws, start_row, entries):
    for offset, (label, formula, fmt) in enumerate(entries):
        label_cell = ws.cell(row=start_row + offset, column=1, value=label)
        label_cell.font = SUMMARY_FONT
        label_cell.fill = SUMMARY_FILL
        value_cell = ws.cell(row=start_row + offset, column=2, value=formula)
        value_cell.font = SUMMARY_FONT
        value_cell.fill = SUMMARY_FILL
        if fmt:
            value_cell.number_format = fmt


def sales_pipeline():
    wb = Workbook()
    ws = wb.active
    ws.title = "Pipeline"
    headers = ["Deal", "Company", "Stage", "Value", "Probability", "Weighted Value"]
    style_sheet(ws, headers, [22, 16, 12, 12, 12, 14], 20)
    rows = [
        ("Website rebuild", "Acme Co", "Proposal", 12000, 0.6),
        ("Annual retainer", "Borealis LLC", "Qualified", 30000, 0.35),
        ("Brand refresh", "Cobalt Studio", "Lead", 8000, 0.15),
        ("Support contract", "Delta Corp", "Won", 15000, 1.0),
        ("Audit project", "Ember Inc", "Lost", 9000, 0.0),
    ]
    for r, (deal, company, stage, value, prob) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=deal)
        ws.cell(row=r, column=2, value=company)
        ws.cell(row=r, column=3, value=stage)
        ws.cell(row=r, column=4, value=value).number_format = "#,##0"
        ws.cell(row=r, column=5, value=prob).number_format = "0%"
        ws.cell(row=r, column=6, value=f"=D{r}*E{r}").number_format = "#,##0"
    summary_block(ws, 8, [
        ("Open pipeline", '=SUMIFS(D2:D6,C2:C6,"<>Won",C2:C6,"<>Lost")', "#,##0"),
        ("Weighted pipeline", "=SUM(F2:F6)", "#,##0"),
        ("Won this period", '=SUMIF(C2:C6,"Won",D2:D6)', "#,##0"),
    ])
    wb.save(OUT / "sales-pipeline-tracker.xlsx")


def budget_tracker():
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    style_sheet(ws, ["Category", "Budget", "Actual", "Variance", "Variance %"], [20, 12, 12, 12, 12], 6)
    rows = [
        ("Rent", 2500, 2500), ("Payroll", 18000, 18450), ("Marketing", 3000, 2100),
        ("Software", 900, 1140), ("Travel", 1200, 480),
    ]
    for r, (cat, budget, actual) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=budget).number_format = "#,##0"
        ws.cell(row=r, column=3, value=actual).number_format = "#,##0"
        ws.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = "#,##0;[Red]-#,##0"
        ws.cell(row=r, column=5, value=f"=IF(B{r}=0,\"\",(C{r}-B{r})/B{r})").number_format = "0.0%"
    summary_block(ws, 8, [
        ("Total budget", "=SUM(B2:B6)", "#,##0"),
        ("Total actual", "=SUM(C2:C6)", "#,##0"),
        ("Total variance", "=SUM(C2:C6)-SUM(B2:B6)", "#,##0;[Red]-#,##0"),
    ])
    wb.save(OUT / "budget-tracker.xlsx")


def invoice_tracker():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    style_sheet(ws, ["Invoice #", "Client", "Issue Date", "Terms (days)", "Due Date", "Amount", "Status", "Days Overdue"],
                [11, 18, 12, 12, 12, 11, 11, 13], 5)
    rows = [
        ("INV-001", "Acme Co", "2026-06-01", 30, 1850, "Paid"),
        ("INV-002", "Borealis LLC", "2026-06-10", 30, 3200, "Sent"),
        ("INV-003", "Cobalt Studio", "2026-06-24", 14, 940, "Sent"),
        ("INV-004", "Acme Co", "2026-07-01", 30, 2100, "Draft"),
    ]
    for r, (num, client, issued, terms, amount, status) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=num)
        ws.cell(row=r, column=2, value=client)
        ws.cell(row=r, column=3, value=date.fromisoformat(issued)).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=4, value=terms)
        ws.cell(row=r, column=5, value=f"=C{r}+D{r}").number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=6, value=amount).number_format = "#,##0"
        ws.cell(row=r, column=7, value=status)
        ws.cell(row=r, column=8, value=f'=IF(G{r}="Paid",0,MAX(0,TODAY()-E{r}))')
    summary_block(ws, 7, [
        ("Outstanding", "=SUMIFS(F2:F5,G2:G5,\"<>Paid\")", "#,##0"),
        ("Overdue invoices", "=COUNTIF(H2:H5,\">0\")", None),
    ])
    wb.save(OUT / "invoice-tracker.xlsx")


def project_task_tracker():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    style_sheet(ws, ["Task", "Owner", "Priority", "Due Date", "Status", "Flag"], [26, 14, 10, 12, 13, 12], 6)
    rows = [
        ("Send contract", "Ana Torres", "High", "2026-06-28", "In Progress"),
        ("Book venue", "Ben Okafor", "Medium", "2026-07-15", "In Progress"),
        ("File permits", "Cara Lim", "High", "2026-07-01", "Complete"),
        ("Draft agenda", "Ana Torres", "Low", "2026-07-20", "Not Started"),
        ("Order badges", "Ben Okafor", "Medium", "2026-07-10", "Not Started"),
    ]
    for r, (task, owner, priority, due, status) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=task)
        ws.cell(row=r, column=2, value=owner)
        ws.cell(row=r, column=3, value=priority)
        ws.cell(row=r, column=4, value=date.fromisoformat(due)).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=5, value=status)
        ws.cell(row=r, column=6, value=f'=IF(AND(D{r}<TODAY(),E{r}<>"Complete"),"Overdue","On Track")')
    summary_block(ws, 9, [
        ("% complete", "=COUNTIF(E2:E6,\"Complete\")/COUNTA(E2:E6)", "0%"),
        ("Overdue tasks", "=COUNTIF(F2:F6,\"Overdue\")", None),
    ])
    wb.save(OUT / "project-task-tracker.xlsx")


def job_application_tracker():
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    style_sheet(ws, ["Company", "Role", "Applied", "Stage", "Next Step", "Days Since Applied"],
                [16, 22, 12, 14, 22, 16], 5)
    rows = [
        ("Acme Co", "Operations Analyst", "2026-06-12", "Interview", "Panel on 2026-07-10"),
        ("Borealis LLC", "HR Coordinator", "2026-06-20", "Applied", "Follow up 2026-07-09"),
        ("Cobalt Studio", "Office Manager", "2026-06-25", "Phone Screen", "Await scheduling"),
        ("Delta Corp", "Program Manager", "2026-07-01", "Applied", "—"),
    ]
    for r, (company, role, applied, stage, next_step) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=company)
        ws.cell(row=r, column=2, value=role)
        ws.cell(row=r, column=3, value=date.fromisoformat(applied)).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=4, value=stage)
        ws.cell(row=r, column=5, value=next_step)
        ws.cell(row=r, column=6, value=f"=TODAY()-C{r}")
    summary_block(ws, 7, [
        ("Active applications", "=COUNTA(A2:A5)", None),
        ("In interview stage", "=COUNTIF(D2:D5,\"Interview\")", None),
    ])
    wb.save(OUT / "job-application-tracker.xlsx")


if __name__ == "__main__":
    sales_pipeline()
    budget_tracker()
    invoice_tracker()
    project_task_tracker()
    job_application_tracker()
    for f in sorted(OUT.glob("*.xlsx")):
        print(f.name, f.stat().st_size, "bytes")
